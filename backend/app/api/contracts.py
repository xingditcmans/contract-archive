"""
合同管理 API
处理合同的增删改查和附件操作
"""
import logging
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Request, status
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from typing import List, Optional
from datetime import date
import os
import uuid
import openpyxl
from io import BytesIO

from app.core.database import get_db
from app.core.security import get_current_user, get_current_admin
from app.core.config import settings
from app.models.database import Contract, Attachment, User, UserRole, Company, ContractStatus
from app.schemas.contract import (
    ContractCreate, ContractUpdate, ContractResponse,
    ContractListResponse, ContractMemoUpdate, ContractStatusUpdate, AttachmentResponse,
    CompanyCreate, CompanyResponse, OCRResult
)
from app.utils.ocr import extract_text_from_pdf, extract_text_from_image, extract_contract_info, diagnose_pdf_extraction, HAS_PYMUPDF, HAS_TESSERACT
from app.utils.ai_extractor import ai_extract_contract_info, get_ai_config
from app.utils.file_security import safe_save_file, sanitize_filename


router = APIRouter(prefix="/api/contracts", tags=["合同管理"])
logger = logging.getLogger("contracts")
audit_logger = logging.getLogger("audit")


def _safe_detail(msg: str) -> str:
    """生产环境返回通用错误消息，防止泄露内部异常详情"""
    if settings.is_production:
        return "操作失败，请稍后重试"
    return msg


# ==================== 公司管理 ====================
@router.get("/companies", response_model=List[CompanyResponse])
async def list_companies(db: Session = Depends(get_db)):
    """获取我方公司列表"""
    companies = db.query(Company).filter(Company.is_active == True).order_by(Company.sort_order).all()
    return companies


@router.post("/companies", response_model=CompanyResponse)
async def create_company(
    company_data: CompanyCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    """创建新公司（仅管理员）"""
    # 检查是否已存在
    existing = db.query(Company).filter(Company.name == company_data.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="公司名称已存在")

    company = Company(**company_data.model_dump())
    db.add(company)
    db.commit()
    db.refresh(company)

    audit_logger.info({
        "action": "COMPANY_CREATED",
        "company_id": company.id,
        "company_name": company_data.name,
        "operator": current_user.username,
        "ip": request.client.host if request.client else "unknown",
    })

    return company


@router.put("/companies/{company_id}", response_model=CompanyResponse)
async def update_company(
    company_id: int,
    company_data: CompanyCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    """更新公司信息（仅管理员）"""
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="公司不存在")

    old_name = company.name
    company.name = company_data.name
    company.keywords = company_data.keywords
    db.commit()
    db.refresh(company)

    audit_logger.info({
        "action": "COMPANY_UPDATED",
        "company_id": company.id,
        "old_name": old_name,
        "new_name": company_data.name,
        "operator": current_user.username,
        "ip": request.client.host if request.client else "unknown",
    })

    return company


@router.delete("/companies/{company_id}")
async def delete_company(
    company_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    """删除公司（仅管理员）"""
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="公司不存在")

    # 软删除
    company.is_active = False
    db.commit()

    audit_logger.warning({
        "action": "COMPANY_DELETED",
        "company_id": company.id,
        "company_name": company.name,
        "operator": current_user.username,
        "ip": request.client.host if request.client else "unknown",
    })

    return {"message": "删除成功"}


# ==================== 合同 CRUD ====================
@router.post("", response_model=ContractResponse)
async def create_contract(
    contract_data: ContractCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """创建新合同"""
    # 检查合同编号唯一性
    existing = db.query(Contract).filter(Contract.contract_no == contract_data.contract_no).first()
    if existing:
        raise HTTPException(status_code=400, detail="合同编号已存在")

    contract = Contract(
        **contract_data.model_dump(),
        created_by=current_user.id
    )
    db.add(contract)
    db.commit()
    db.refresh(contract)

    audit_logger.info({
        "action": "CONTRACT_CREATED",
        "contract_id": contract.id,
        "contract_no": contract_data.contract_no,
        "contract_name": contract_data.contract_name,
        "operator": current_user.username,
        "ip": request.client.host if request.client else "unknown",
    })

    # 加载关联数据
    return await _load_contract_response(db, contract)


@router.get("")
async def list_contracts(
    keyword: Optional[str] = Query(None, description="全局关键词搜索"),
    contract_no: Optional[str] = Query(None, description="合同编号搜索"),
    contract_name: Optional[str] = Query(None, description="合同名称搜索"),
    counterparty: Optional[str] = Query(None, description="对方公司搜索"),
    contract_type: Optional[str] = Query(None, description="合同类型"),
    company_id: Optional[int] = Query(None, description="我方公司"),
    department: Optional[str] = Query(None, description="部门"),
    amount: Optional[str] = Query(None, description="金额搜索"),
    submit_date_from: Optional[date] = Query(None, description="提交日期起"),
    submit_date_to: Optional[date] = Query(None, description="提交日期止"),
    amount_min: Optional[float] = Query(None, description="最小金额"),
    amount_max: Optional[float] = Query(None, description="最大金额"),
    sort_by: Optional[str] = Query("created_at", description="排序字段"),
    sort_order: Optional[str] = Query("desc", description="排序方向: asc/desc"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    db: Session = Depends(get_db)
):
    """查询合同列表（支持多条件筛选、排序和分页）"""
    query = db.query(Contract).join(User, Contract.created_by == User.id)

    # 全局关键词搜索（模糊匹配所有主要字段，包括金额和备注）
    if keyword:
        keyword_filter = f"%{keyword}%"
        query = query.filter(
            or_(
                Contract.contract_no.ilike(keyword_filter),
                Contract.contract_name.ilike(keyword_filter),
                Contract.counterparty.ilike(keyword_filter),
                Contract.department.ilike(keyword_filter),
                Contract.handler.ilike(keyword_filter),
                Contract.remarks.ilike(keyword_filter),
                Contract.memo.ilike(keyword_filter),
                Contract.amount.ilike(keyword_filter),
                Contract.paper_copies.ilike(keyword_filter),
            )
        )

    # 独立字段搜索（与全局搜索互补）
    if contract_no:
        query = query.filter(Contract.contract_no.ilike(f"%{contract_no}%"))
    if contract_name:
        query = query.filter(Contract.contract_name.ilike(f"%{contract_name}%"))
    if counterparty:
        query = query.filter(Contract.counterparty.ilike(f"%{counterparty}%"))
    if amount:
        query = query.filter(Contract.amount.ilike(f"%{amount}%"))

    # 条件筛选
    if contract_type:
        query = query.filter(Contract.contract_type == contract_type)
    if company_id:
        query = query.filter(Contract.company_id == company_id)
    if department:
        query = query.filter(Contract.department.ilike(f"%{department}%"))
    if submit_date_from:
        query = query.filter(Contract.submit_date >= submit_date_from)
    if submit_date_to:
        query = query.filter(Contract.submit_date <= submit_date_to)

    # 排序（安全校验：只允许指定字段）
    allowed_sort_fields = {
        "contract_no": Contract.contract_no,
        "contract_name": Contract.contract_name,
        "contract_type": Contract.contract_type,
        "counterparty": Contract.counterparty,
        "submit_date": Contract.submit_date,
        "amount": Contract.amount,
        "department": Contract.department,
        "created_at": Contract.created_at,
    }
    sort_column = allowed_sort_fields.get(sort_by, Contract.created_at)
    if sort_order == "asc":
        query = query.order_by(sort_column.asc())
    else:
        query = query.order_by(sort_column.desc())

    # 分页
    total = query.count()
    contracts = query.offset((page - 1) * page_size).limit(page_size).all()

    # 批量查询附件计数（避免 N+1）
    contract_ids = [c.id for c in contracts]
    att_counts = {}
    if contract_ids:
        from sqlalchemy import func as sqlfunc
        count_rows = (
            db.query(Attachment.contract_id, sqlfunc.count(Attachment.id))
            .filter(Attachment.contract_id.in_(contract_ids))
            .group_by(Attachment.contract_id)
            .all()
        )
        att_counts = {cid: cnt for cid, cnt in count_rows}

    # 转换为响应格式
    result = []
    for c in contracts:
        result.append(ContractListResponse(
            id=c.id,
            contract_no=c.contract_no,
            contract_name=c.contract_name,
            contract_type=c.contract_type.value,
            company_name=c.company.name if c.company else None,
            counterparty=c.counterparty,
            submit_date=c.submit_date,
            amount=c.amount,
            department=c.department,
            handler=c.handler,
            memo=c.memo,
            attachment_count=att_counts.get(c.id, 0),
            created_at=c.created_at,
            creator_name=c.creator.username if c.creator else None,
            status=c.status.value if c.status else "有效"
        ))

    return {
        "items": result,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.get("/export")
async def export_contracts(
    keyword: Optional[str] = Query(None),
    contract_type: Optional[str] = Query(None),
    company_id: Optional[int] = Query(None),
    counterparty: Optional[str] = Query(None),
    department: Optional[str] = Query(None),
    submit_date_from: Optional[date] = Query(None),
    submit_date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """导出合同列表为 Excel"""
    # 复用查询逻辑
    query = db.query(Contract).join(User, Contract.created_by == User.id).join(Company, Contract.company_id == Company.id, isouter=True)

    if keyword:
        keyword_filter = f"%{keyword}%"
        query = query.filter(
            or_(
                Contract.contract_no.ilike(keyword_filter),
                Contract.contract_name.ilike(keyword_filter),
                Contract.counterparty.ilike(keyword_filter)
            )
        )
    if contract_type:
        query = query.filter(Contract.contract_type == contract_type)
    if company_id:
        query = query.filter(Contract.company_id == company_id)
    if counterparty:
        query = query.filter(Contract.counterparty.ilike(f"%{counterparty}%"))
    if department:
        query = query.filter(Contract.department.ilike(f"%{department}%"))
    if submit_date_from:
        query = query.filter(Contract.submit_date >= submit_date_from)
    if submit_date_to:
        query = query.filter(Contract.submit_date <= submit_date_to)

    contracts = query.order_by(Contract.created_at.desc()).all()

    # 创建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合同列表"

    # 表头
    headers = ["合同编号", "合同名称", "合同类型", "我方公司", "对方公司", "提交日期",
               "有效期至", "部门", "经办人", "金额", "纸质份数", "备注", "备忘录", "录入时间", "录入人"]
    ws.append(headers)

    # 数据行
    for c in contracts:
        ws.append([
            c.contract_no, c.contract_name, c.contract_type.value,
            c.company.name if c.company else "",
            c.counterparty, str(c.submit_date),
            str(c.valid_until) if c.valid_until else "",
            c.department, c.handler or "", c.amount or "",
            c.paper_copies or "", c.remarks or "", c.memo or "",
            c.created_at.strftime("%Y-%m-%d %H:%M"),
            c.creator.username if c.creator else ""
        ])

    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # 返回文件
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contracts_export.xlsx"}
    )


@router.get("/{contract_id}", response_model=ContractResponse)
async def get_contract(
    contract_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取合同详情"""
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")

    return await _load_contract_response(db, contract)


@router.put("/{contract_id}", response_model=ContractResponse)
async def update_contract(
    contract_id: int,
    contract_data: ContractUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """更新合同（普通用户只能修改自己的合同）"""
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")

    # 权限检查
    if current_user.role != UserRole.ADMIN and contract.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="只能修改自己创建的合同")

    # 更新字段
    update_data = contract_data.model_dump(exclude_unset=True)

    # 检查合同编号唯一性（修改编号时不能与其他合同重复）
    if "contract_no" in update_data and update_data["contract_no"] != contract.contract_no:
        existing = db.query(Contract).filter(
            Contract.contract_no == update_data["contract_no"],
            Contract.id != contract_id
        ).first()
        if existing:
            raise HTTPException(
                status_code=400,
                detail=f"合同编号 {update_data['contract_no']} 已被合同「{existing.contract_name}」（对方: {existing.counterparty}）使用，请更换编号"
            )

    for key, value in update_data.items():
        setattr(contract, key, value)

    db.commit()
    db.refresh(contract)

    audit_logger.info({
        "action": "CONTRACT_UPDATED",
        "contract_id": contract.id,
        "contract_no": contract.contract_no,
        "changed_fields": list(update_data.keys()),
        "operator": current_user.username,
        "ip": request.client.host if request.client else "unknown",
    })

    return await _load_contract_response(db, contract)


@router.put("/{contract_id}/memo", response_model=ContractResponse)
async def update_memo(
    contract_id: int,
    memo_data: ContractMemoUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """更新备忘录（任何人都可以随时修改）"""
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")

    contract.memo = memo_data.memo
    db.commit()
    db.refresh(contract)

    return await _load_contract_response(db, contract)


@router.put("/{contract_id}/status")
async def update_contract_status(
    contract_id: int,
    status_data: ContractStatusUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    """更新合同状态（仅管理员：有效/作废）"""
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")

    old_status = contract.status.value if contract.status else "有效"
    contract.status = ContractStatus(status_data.status)
    db.commit()
    db.refresh(contract)

    audit_logger.info({
        "action": "CONTRACT_STATUS_CHANGED",
        "contract_id": contract.id,
        "contract_no": contract.contract_no,
        "old_status": old_status,
        "new_status": contract.status.value,
        "operator": current_user.username,
        "ip": request.client.host if request.client else "unknown",
    })

    return {"id": contract.id, "status": contract.status.value}


@router.delete("/{contract_id}")
async def delete_contract(
    contract_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """删除合同（仅管理员）"""
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="需要管理员权限")

    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")

    # 记录审计日志（在删除前）
    contract_no = contract.contract_no
    contract_name = contract.contract_name

    # 删除附件文件
    for attachment in contract.attachments:
        try:
            if os.path.exists(attachment.file_path):
                os.remove(attachment.file_path)
        except Exception:
            pass

    db.delete(contract)
    db.commit()

    audit_logger.warning({
        "action": "CONTRACT_DELETED",
        "contract_id": contract_id,
        "contract_no": contract_no,
        "contract_name": contract_name,
        "operator": current_user.username,
        "ip": request.client.host if request.client else "unknown",
    })

    return {"message": "删除成功"}


# ==================== 附件管理 ====================
@router.get("/{contract_id}/attachments", response_model=List[AttachmentResponse])
async def list_attachments(
    contract_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取合同附件列表（仅 id/file_name/file_type，供列表页预览用）"""
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")

    attachments = db.query(Attachment).filter(Attachment.contract_id == contract_id).all()
    return [AttachmentResponse(
        id=a.id,
        file_name=a.file_name,
        file_type=a.file_type,
        file_size=a.file_size,
        uploaded_at=a.uploaded_at,
        uploaded_by=a.uploaded_by,
        uploader_name=current_user.username
    ) for a in attachments]


@router.post("/{contract_id}/attachments")
async def upload_attachment(
    contract_id: int,
    request: Request,
    files: List[UploadFile] = File(...),
    replace_existing: bool = Query(False, description="是否替换同名附件（兼容旧参数）"),
    mode: Optional[str] = Query(None, description="冲突处理模式: skip | replace | add_new"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """上传合同附件（支持多个）
    
    冲突处理模式（mode 参数，优先级高于 replace_existing）：
    - skip: 跳过同名文件，只上传不冲突的
    - replace: 替换同名文件（同 replace_existing=true）
    - add_new: 保留旧文件，新增文件自动重命名（如 合同.pdf → 合同_1.pdf）
    - 不传 mode 且不传 replace_existing: 返回 409 让前端弹窗选择
    """
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")

    # 权限检查
    if current_user.role != UserRole.ADMIN and contract.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="只能为自己的合同上传附件")

    # 先检测同名附件冲突（在上传前全部检测，避免部分写入）
    conflicts = []
    conflict_names = set()
    for file in files:
        existing_attachment = db.query(Attachment).filter(
            Attachment.contract_id == contract_id,
            Attachment.file_name == file.filename
        ).first()
        if existing_attachment:
            conflicts.append({
                "filename": file.filename,
                "new_size": None,
                "existing_id": existing_attachment.id,
                "existing_size": existing_attachment.file_size,
                "existing_uploaded_at": existing_attachment.uploaded_at.isoformat() if existing_attachment.uploaded_at else None,
            })
            conflict_names.add(file.filename)

    # 决定冲突处理策略
    effective_mode = mode or ("replace" if replace_existing else None)

    if conflicts and not effective_mode:
        # 无模式 → 返回409让前端弹窗选择
        from fastapi.responses import JSONResponse
        return JSONResponse(
            status_code=409,
            content={
                "detail": f"以下 {len(conflicts)} 个文件已存在同名附件，请选择处理方式",
                "conflicts": conflicts,
                "available_modes": ["skip", "replace", "add_new"],
            }
        )

    uploaded = []
    skipped = []

    for file in files:
        is_conflict = file.filename in conflict_names

        # === 模式: skip — 跳过冲突文件 ===
        if effective_mode == "skip" and is_conflict:
            skipped.append({"filename": file.filename, "reason": "skip"})
            continue

        # === 模式: replace — 先删旧再写新 ===
        if effective_mode == "replace" and is_conflict:
            existing_attachment = db.query(Attachment).filter(
                Attachment.contract_id == contract_id,
                Attachment.file_name == file.filename
            ).first()
            if existing_attachment:
                try:
                    if os.path.exists(existing_attachment.file_path):
                        os.remove(existing_attachment.file_path)
                except Exception:
                    pass
                db.delete(existing_attachment)

        # === 模式: add_new — 自动重命名 ===
        if effective_mode == "add_new" and is_conflict:
            base, ext = os.path.splitext(file.filename)
            counter = 1
            while True:
                new_name = f"{base}_{counter}{ext}"
                existing = db.query(Attachment).filter(
                    Attachment.contract_id == contract_id,
                    Attachment.file_name == new_name
                ).first()
                if not existing:
                    break
                counter += 1
            file.filename = new_name
            # 更新 conflicts 中的文件名用于返回
            for c in conflicts:
                if c["filename"] == f"{base}{ext}":
                    c["renamed_to"] = new_name

        # 读取文件内容
        content = await file.read()

        # 更新 conflicts 中的文件大小
        if conflicts:
            for c in conflicts:
                if c["filename"] == file.filename or c.get("renamed_to") == file.filename:
                    c["new_size"] = len(content)

        # 验证文件类型（通过扩展名）
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in settings.ALLOWED_EXTENSIONS:
            raise HTTPException(status_code=400, detail=f"不支持的文件类型: {ext}")

        # 使用安全保存函数（包含内容验证）
        try:
            file_path, original_name, file_hash = safe_save_file(
                content=content,
                upload_dir=settings.UPLOAD_DIR,
                original_filename=file.filename,
                max_size=settings.MAX_FILE_SIZE,
                contract_no=contract.contract_no
            )
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=_safe_detail(f"文件保存失败: {str(e)}"))

        # 从路径中提取存储的文件名
        stored_name = os.path.basename(file_path)
        if ext in [".pdf"]:
            file_type = "pdf"
        elif ext in [".jpg", ".jpeg", ".png"]:
            file_type = "image"
        else:
            file_type = "word"

        # 保存到数据库
        attachment = Attachment(
            contract_id=contract_id,
            file_name=original_name,  # 保存清理后的原始文件名
            stored_name=stored_name,
            file_path=file_path,
            file_type=file_type,
            file_size=len(content),
            uploaded_by=current_user.id
        )
        db.add(attachment)
        uploaded.append(attachment)

    db.commit()

    # ===== 附件自动重命名：改为"合同编号 合同名称[序号].ext" =====
    # 重命名策略：
    # - 该合同第1个附件: "合同编号 合同名称.ext"
    # - 第2个起: "合同编号 合同名称1.ext", "合同编号 合同名称2.ext", ...
    # - 序号基于全部已存在附件数（含本次上传前的），保证连续不重复
    def _sanitize_for_filename(s: str) -> str:
        """去除文件名中不合法的字符"""
        import re
        return re.sub(r'[\\/:*?"<>|]', '_', s).strip()

    for a in uploaded:
        db.refresh(a)
        ext = os.path.splitext(a.file_name)[1].lower()
        base = f"{_sanitize_for_filename(contract.contract_no)} {_sanitize_for_filename(contract.contract_name)}"

        # 该合同已存在多少个附件（含本次已写入的）
        total_att = db.query(Attachment).filter(Attachment.contract_id == contract_id).count()
        # 当前这条是第几个（按 id 升序排列）
        att_rank = (
            db.query(Attachment)
            .filter(Attachment.contract_id == contract_id, Attachment.id <= a.id)
            .count()
        )

        # 第1个附件不加数字后缀，第2个起加 1、2、3...
        if att_rank == 1:
            new_display_name = f"{base}{ext}"
        else:
            new_display_name = f"{base}{att_rank - 1}{ext}"

        # 检查目标名称是否已被其他附件占用（避免本次批量上传多个互相冲突）
        conflict = db.query(Attachment).filter(
            Attachment.contract_id == contract_id,
            Attachment.file_name == new_display_name,
            Attachment.id != a.id
        ).first()
        if conflict:
            # 找下一个可用序号
            seq = att_rank
            while True:
                candidate = f"{base}{seq}{ext}"
                if not db.query(Attachment).filter(
                    Attachment.contract_id == contract_id,
                    Attachment.file_name == candidate,
                    Attachment.id != a.id
                ).first():
                    new_display_name = candidate
                    break
                seq += 1

        a.file_name = new_display_name

    db.commit()

    # 审计日志
    audit_logger.info({
        "action": "ATTACHMENT_UPLOADED",
        "contract_id": contract_id,
        "file_count": len(uploaded),
        "skipped_count": len(skipped),
        "file_names": [a.file_name for a in uploaded],
        "mode": effective_mode,
        "operator": current_user.username,
        "ip": request.client.host if request.client else "unknown",
    })

    # 返回上传结果
    result = []
    for a in uploaded:
        db.refresh(a)
        result.append(AttachmentResponse(
            id=a.id,
            file_name=a.file_name,
            file_type=a.file_type,
            file_size=a.file_size,
            uploaded_at=a.uploaded_at,
            uploaded_by=a.uploaded_by,
            uploader_name=current_user.username
        ))

    response_data = {
        "attachments": result,
        "skipped": skipped,
        "mode_used": effective_mode,
    }
    return response_data


@router.get("/{contract_id}/attachments/{attachment_id}/download")
async def download_attachment(
    contract_id: int,
    attachment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """下载附件"""
    attachment = db.query(Attachment).filter(
        Attachment.id == attachment_id,
        Attachment.contract_id == contract_id
    ).first()

    if not attachment:
        raise HTTPException(status_code=404, detail="附件不存在")

    if not os.path.exists(attachment.file_path):
        raise HTTPException(status_code=404, detail="文件已丢失")

    return FileResponse(
        attachment.file_path,
        filename=attachment.file_name,
        media_type="application/octet-stream"
    )


@router.get("/{contract_id}/attachments/{attachment_id}/preview")
async def preview_attachment(
    contract_id: int,
    attachment_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """预览附件（返回适合浏览器内嵌查看的MIME类型）
    
    注意：预览不要求 auth header，因为 iframe/img 请求无法带 Authorization。
    改为从 query string 中验证 token。
    """
    # 从 query string 验证 token（iframe/img 请求无法带 Authorization header）
    token = request.query_params.get("token")
    if token:
        from app.core.security import decode_token
        payload = decode_token(token)
        if payload is None:
            raise HTTPException(status_code=401, detail="登录已过期，请重新登录")
    else:
        raise HTTPException(status_code=401, detail="缺少认证token")

    attachment = db.query(Attachment).filter(
        Attachment.id == attachment_id,
        Attachment.contract_id == contract_id
    ).first()

    if not attachment:
        raise HTTPException(status_code=404, detail="附件不存在")

    if not os.path.exists(attachment.file_path):
        raise HTTPException(status_code=404, detail="文件已丢失")

    # 根据文件类型返回正确的 MIME type
    mime_map = {
        "pdf": "application/pdf",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "png": "image/png",
        "gif": "image/gif",
        "bmp": "image/bmp",
        "webp": "image/webp",
        "doc": "application/msword",
        "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    }
    media_type = mime_map.get(attachment.file_type, "application/octet-stream")

    # Content-Disposition: inline 让浏览器预览；用 RFC 5987 编码处理中文文件名
    from urllib.parse import quote
    encoded_filename = quote(attachment.file_name, safe="")
    return FileResponse(
        attachment.file_path,
        media_type=media_type,
        headers={
            "Content-Disposition": f"inline; filename*=UTF-8''{encoded_filename}"
        },
    )


@router.delete("/{contract_id}/attachments/{attachment_id}")
async def delete_attachment(
    contract_id: int,
    attachment_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """删除附件"""
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="需要管理员权限")

    attachment = db.query(Attachment).filter(
        Attachment.id == attachment_id,
        Attachment.contract_id == contract_id
    ).first()

    if not attachment:
        raise HTTPException(status_code=404, detail="附件不存在")

    # 记录信息用于审计
    file_name = attachment.file_name
    file_path = attachment.file_path

    # 删除文件
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except Exception:
        pass

    db.delete(attachment)
    db.commit()

    audit_logger.warning({
        "action": "ATTACHMENT_DELETED",
        "contract_id": contract_id,
        "attachment_id": attachment_id,
        "file_name": file_name,
        "operator": current_user.username,
        "ip": request.client.host if request.client else "unknown",
    })

    return {"message": "删除成功"}


# ==================== OCR 识别 ====================
@router.post("/ocr", response_model=OCRResult)
async def ocr_recognize(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """OCR 识别 PDF 或图片中的文字（异步执行，防止阻塞事件循环）"""
    import asyncio
    import traceback

    # 读取文件内容
    content = await file.read()
    ext = os.path.splitext(file.filename)[1].lower()

    if ext not in [".pdf", ".jpg", ".jpeg", ".png"]:
        raise HTTPException(status_code=400, detail="仅支持 PDF 或图片文件")

    # ⚠️ OCR/Tesseract 是 CPU 密集型操作，必须在独立线程中执行
    #    否则会阻塞 uvicorn 事件循环，导致前端超时/服务无响应
    loop = asyncio.get_event_loop()
    diagnostics = None

    try:
        if ext == ".pdf":
            # PDF：文本提取 + 诊断（均在独立线程运行）
            raw_text = await loop.run_in_executor(None, extract_text_from_pdf, content)
            diagnostics = await loop.run_in_executor(None, diagnose_pdf_extraction, content)
        else:
            # 图片：直接 OCR
            raw_text = await loop.run_in_executor(None, extract_text_from_image, content)

    except Exception as e:
        logger.error(f"OCR 处理异常: {e}\n{traceback.format_exc()}")
        # 返回失败结果而非抛异常，让前端能拿到诊断信息
        return OCRResult(
            raw_text=f"[处理失败] {type(e).__name__}: {str(e)[:500]}",
            diagnostics={
                "engines": {"pdfplumber": True, "pymupdf": HAS_PYMUPDF, "tesseract": HAS_TESSERACT},
                "error": f"{type(e).__name__}: {str(e)[:500]}",
                "result": "exception",
            },
        )

    # 读取我方公司列表（用于双向匹配）
    companies = db.query(Company).filter(Company.is_active == True).order_by(Company.sort_order).all()
    known_companies = [{"id": c.id, "name": c.name, "keywords": c.keywords} for c in companies]

    ai_config = get_ai_config()
    ai_enabled = ai_config.get("enabled", False)
    # fallback_only=True  → 兜底模式：正则优先，正则搞不定的字段再让 AI 补
    # fallback_only=False → 全量模式：直接跳过正则，AI 全权负责（快且准）
    fallback_only = ai_config.get("fallback_only", True)

    ai_extracted = None
    extracted = {}

    if ai_enabled and not fallback_only:
        # ===== 全量 AI 模式：跳过正则，直接给 AI =====
        logger.info("全量 AI 模式：跳过正则，直接调用 AI")
        try:
            company_names = [c["name"] for c in known_companies]
            ai_extracted = await ai_extract_contract_info(
                raw_text,
                known_companies=company_names,
                ocr_result=None,
            )
            if ai_extracted:
                logger.info(f"AI 提取字段: {[k for k in ai_extracted if not k.startswith('_')]}")
        except Exception as e:
            logger.warning(f"AI 提取调用失败，降级到正则: {e}")
            extracted = extract_contract_info(raw_text, known_companies=known_companies)
    else:
        # ===== 正则优先模式（纯正则 或 兜底 AI）=====
        extracted = extract_contract_info(raw_text, known_companies=known_companies)

        if ai_enabled and fallback_only:
            # 兜底：先看正则缺了哪些关键字段，缺了再调 AI
            missing_fields = [
                k for k in ["contract_name", "counterparty", "amount", "date"]
                if not extracted.get(k)
            ]
            if missing_fields:
                logger.info(f"兜底 AI 模式：正则缺少 {missing_fields}，调用 AI 补充")
                try:
                    company_names = [c["name"] for c in known_companies]
                    ai_extracted = await ai_extract_contract_info(
                        raw_text,
                        known_companies=company_names,
                        ocr_result=extracted,
                    )
                    if ai_extracted:
                        logger.info(f"AI 补充字段: {[k for k in ai_extracted if not k.startswith('_')]}")
                except Exception as e:
                    logger.warning(f"兜底 AI 调用失败（不影响正则结果）: {e}")
            else:
                logger.info("兜底 AI 模式：正则已提取所有关键字段，跳过 AI 调用")

    # 将 AI 结果映射到标准 extracted 字典
    if ai_extracted:
        ai_field_map = {
            "contract_name": "contract_name",
            "counterparty": "counterparty",
            "amount": "amount",
            "sign_date": "date",
            "date": "date",
        }
        for ai_key, ocr_key in ai_field_map.items():
            if ai_extracted.get(ai_key):
                if not fallback_only:
                    # 全量模式：AI 无条件覆盖
                    extracted[ocr_key] = ai_extracted[ai_key]
                else:
                    # 兜底模式：只填正则没拿到的字段
                    if not extracted.get(ocr_key):
                        extracted[ocr_key] = ai_extracted[ai_key]
        # AI 补充我方公司匹配
        if ai_extracted.get("our_company"):
            ai_our = ai_extracted["our_company"]
            for c in known_companies:
                if c["name"] in ai_our or ai_our in c["name"]:
                    extracted["matched_company_id"] = c["id"]
                    extracted["matched_company_name"] = c["name"]
                    break
        extracted["_ai_enhanced"] = True

    # 确定最终 AI 模式标识
    if not ai_enabled:
        _ai_mode = "none"
    elif not fallback_only:
        _ai_mode = "full"
    else:
        _ai_mode = "fallback" if ai_extracted else "none"

    # 返回识别结果和原始文本供人工核对
    return OCRResult(
        contract_no=extracted.get("contract_no"),
        contract_name=extracted.get("contract_name"),
        counterparty=extracted.get("counterparty"),
        amount=extracted.get("amount"),
        date=extracted.get("date"),
        raw_text=raw_text,
        matched_company_id=extracted.get("matched_company_id"),
        matched_company_name=extracted.get("matched_company_name"),
        counterparty_candidates=extracted.get("counterparty_candidates"),
        diagnostics=diagnostics,
        ai_enhanced=extracted.get("_ai_enhanced", False),
        ai_raw=ai_extracted,
        ai_mode=_ai_mode,  # v3.5: full / fallback / none
    )


# ==================== 批量导入 ====================
@router.post("/batch-import")
async def batch_import_contracts(
    request: Request,
    excel_file: UploadFile = File(...),
    pdf_files: List[UploadFile] = File(default=[]),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    批量导入合同
    上传 Excel 文件（含合同信息）+ 可选的 PDF 附件

    Excel 格式要求：
    - 第一行为表头（自动识别列名）
    - 支持列名：合同编号、合同名称、合同类型、对方公司、提交日期、部门、金额 等
    - 列名模糊匹配，支持多种命名习惯

    PDF 附件匹配规则：
    - 按合同编号匹配（文件名包含合同编号即关联）
    - 未匹配的 PDF 列在返回结果的 unmatched_pdfs 中
    """
    # 读取 Excel
    excel_content = await excel_file.read()
    wb = openpyxl.load_workbook(BytesIO(excel_content))
    ws = wb.active

    # 读取表头
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value else "")

    # 列名映射表（支持多种命名习惯）
    column_map = {
        "contract_no": ["合同编号", "编号", "合同号", "审批编号", "云之家编号", "contract_no", "no"],
        "contract_name": ["合同名称", "名称", "项目名称", "标题", "contract_name", "name", "title"],
        "contract_type": ["合同类型", "类型", "type", "contract_type"],
        "counterparty": ["对方公司", "对方", "乙方", "供应商", "客户", "对方单位", "counterparty"],
        "submit_date": ["提交日期", "日期", "签订日期", "签约日期", "date", "submit_date"],
        "department": ["部门", "所属部门", "申请部门", "department"],
        "handler": ["经办人", "负责人", "handler"],
        "amount": ["金额", "合同金额", "总价", "amount"],
        "paper_copies": ["份数", "纸质份数", "copies"],
        "remarks": ["备注", "说明", "remarks"],
    }

    # 映射列索引
    col_index = {}
    for field, aliases in column_map.items():
        for alias in aliases:
            for idx, header in enumerate(headers):
                if alias.lower() in header.lower():
                    col_index[field] = idx
                    break
            if field in col_index:
                break

    # 解析数据行
    result = {
        "total": 0,
        "success": 0,
        "skipped": 0,
        "errors": [],        # 结构化错误：[{row, type, message, ...}]
        "unmatched_pdfs": [],
    }

    # 处理 PDF 文件名映射
    pdf_name_map = {}
    for pdf in pdf_files:
        pdf_name_map[pdf.filename] = pdf

    matched_pdfs = set()
    imported_contracts = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        result["total"] += 1

        # 跳过空行
        if not any(row):
            result["skipped"] += 1
            continue

        # 提取字段
        contract_data = {}
        for field, idx in col_index.items():
            if idx < len(row) and row[idx] is not None:
                val = str(row[idx]).strip()
                if val and val.lower() not in ("none", "null", ""):
                    contract_data[field] = val

        # 必要字段检查
        if "contract_no" not in contract_data:
            result["errors"].append({
                "row": row_idx,
                "type": "missing_field",
                "field": "contract_no",
                "message": f"第{row_idx}行：缺少合同编号"
            })
            result["skipped"] += 1
            continue
        if "contract_name" not in contract_data:
            result["errors"].append({
                "row": row_idx,
                "type": "missing_field",
                "field": "contract_name",
                "message": f"第{row_idx}行：缺少合同名称"
            })
            result["skipped"] += 1
            continue
        if "counterparty" not in contract_data:
            contract_data["counterparty"] = "未知"

        # 检查合同编号是否与 Excel 内前面的行重复
        already_in_batch = [c for c in imported_contracts if c == contract_data["contract_no"]]
        if already_in_batch:
            result["errors"].append({
                "row": row_idx,
                "type": "contract_no_duplicate",
                "contract_no": contract_data["contract_no"],
                "duplicate_source": "本批次内重复",
                "message": f"第{row_idx}行：合同编号 {contract_data['contract_no']} 与前面已导入的行重复，跳过"
            })
            result["skipped"] += 1
            continue

        # 检查合同编号是否已在数据库中存在
        existing = db.query(Contract).filter(
            Contract.contract_no == contract_data["contract_no"]
        ).first()
        if existing:
            result["errors"].append({
                "row": row_idx,
                "type": "contract_no_duplicate",
                "contract_no": contract_data["contract_no"],
                "duplicate_source": "数据库已有记录",
                "existing_id": existing.id,
                "existing_name": existing.contract_name,
                "existing_counterparty": existing.counterparty,
                "existing_submit_date": str(existing.submit_date) if existing.submit_date else "",
                "existing_department": existing.department or "",
                "message": f"第{row_idx}行：合同编号 {contract_data['contract_no']} 已存在（ID: {existing.id}，名称: {existing.contract_name}，对方: {existing.counterparty}），跳过"
            })
            result["skipped"] += 1
            continue

        # 日期格式处理
        submit_date = None
        if "submit_date" in contract_data:
            for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日", "%m/%d/%Y"]:
                try:
                    from datetime import datetime
                    submit_date = datetime.strptime(contract_data["submit_date"], fmt).date()
                    break
                except ValueError:
                    continue

        # 合同类型标准化
        contract_type_val = "其他"
        if "contract_type" in contract_data:
            ct = contract_data["contract_type"]
            if any(kw in ct for kw in ["采购", "购买", "购入"]):
                contract_type_val = "采购"
            elif any(kw in ct for kw in ["销售", "出售", "卖出"]):
                contract_type_val = "销售"

        # 创建合同
        try:
            contract = Contract(
                contract_no=contract_data["contract_no"],
                contract_name=contract_data.get("contract_name", ""),
                contract_type=contract_type_val,
                counterparty=contract_data.get("counterparty", "未知"),
                submit_date=submit_date or date.today(),
                department=contract_data.get("department", "未知部门"),
                handler=contract_data.get("handler"),
                amount=contract_data.get("amount"),
                paper_copies=contract_data.get("paper_copies"),
                remarks=contract_data.get("remarks"),
                created_by=current_user.id
            )
            db.add(contract)
            db.flush()  # 获取 contract.id

            # 匹配 PDF 附件
            contract_no = contract_data["contract_no"]
            for pdf_name, pdf_file in pdf_name_map.items():
                if pdf_name in matched_pdfs:
                    continue
                # 文件名中包含合同编号即匹配
                if contract_no in pdf_name or pdf_name.replace(".pdf", "") in contract_no:
                    # 检查此合同是否已有同名附件（数据库级重复检测）
                    existing_attachment = db.query(Attachment).filter(
                        Attachment.contract_id == contract.id,
                        Attachment.file_name == pdf_name
                    ).first()
                    if existing_attachment:
                        result["errors"].append({
                            "row": row_idx,
                            "type": "attachment_duplicate",
                            "contract_no": contract_no,
                            "pdf_name": pdf_name,
                            "existing_attachment_id": existing_attachment.id,
                            "existing_uploaded_at": existing_attachment.uploaded_at.isoformat() if existing_attachment.uploaded_at else "",
                            "message": f"第{row_idx}行（{contract_no}）：附件 {pdf_name} 已存在（附件ID: {existing_attachment.id}），跳过"
                        })
                        matched_pdfs.add(pdf_name)
                        continue

                    content = await pdf_file.read()
                    await pdf_file.seek(0)  # 重置指针

                    try:
                        file_path, original_name, file_hash = safe_save_file(
                            content=content,
                            upload_dir=settings.UPLOAD_DIR,
                            original_filename=pdf_name,
                            max_size=settings.MAX_FILE_SIZE,
                            contract_no=contract_no
                        )
                        attachment = Attachment(
                            contract_id=contract.id,
                            file_name=original_name,
                            stored_name=os.path.basename(file_path),
                            file_path=file_path,
                            file_type="pdf",
                            file_size=len(content),
                            uploaded_by=current_user.id
                        )
                        db.add(attachment)
                        matched_pdfs.add(pdf_name)
                    except Exception as e:
                        result["errors"].append({
                            "row": row_idx,
                            "type": "attachment_save_failed",
                            "contract_no": contract_no,
                            "pdf_name": pdf_name,
                            "message": _safe_detail(f"合同 {contract_no} 的附件 {pdf_name} 保存失败: {str(e)}")
                        })

            imported_contracts.append(contract_no)
            result["success"] += 1

        except Exception as e:
            result["errors"].append({
                "row": row_idx,
                "type": "create_failed",
                "message": _safe_detail(f"第{row_idx}行：创建失败 - {str(e)}")
            })
            result["skipped"] += 1

    db.commit()

    # 统计未匹配的 PDF
    for pdf_name in pdf_name_map:
        if pdf_name not in matched_pdfs:
            result["unmatched_pdfs"].append(pdf_name)

    audit_logger.warning({
        "action": "BATCH_IMPORT",
        "total": result["total"],
        "success": result["success"],
        "skipped": result["skipped"],
        "operator": current_user.username,
        "ip": request.client.host if request.client else "unknown",
    })

    return result


# ==================== 辅助函数 ====================
async def _load_contract_response(db: Session, contract: Contract) -> ContractResponse:
    """加载合同响应数据（含关联信息）"""
    # 加载创建者
    creator = db.query(User).filter(User.id == contract.created_by).first()

    # 加载附件
    attachments = db.query(Attachment).filter(Attachment.contract_id == contract.id).all()

    # 加载附件上传者
    attachment_responses = []
    for a in attachments:
        uploader = db.query(User).filter(User.id == a.uploaded_by).first()
        attachment_responses.append(AttachmentResponse(
            id=a.id,
            file_name=a.file_name,
            file_type=a.file_type,
            file_size=a.file_size,
            uploaded_at=a.uploaded_at,
            uploaded_by=a.uploaded_by,
            uploader_name=uploader.username if uploader else None
        ))

    return ContractResponse(
        id=contract.id,
        contract_no=contract.contract_no,
        contract_name=contract.contract_name,
        contract_type=contract.contract_type,
        company_id=contract.company_id,
        company_name=contract.company.name if contract.company else None,
        counterparty=contract.counterparty,
        submit_date=contract.submit_date,
        valid_until=contract.valid_until,
        department=contract.department,
        handler=contract.handler,
        amount=contract.amount,
        paper_copies=contract.paper_copies,
        remarks=contract.remarks,
        memo=contract.memo,
        status=contract.status.value if contract.status else "有效",
        created_by=contract.created_by,
        creator_name=creator.username if creator else None,
        created_at=contract.created_at,
        updated_at=contract.updated_at,
        attachments=attachment_responses
    )
